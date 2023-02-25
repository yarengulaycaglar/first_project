import openpyxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

def all_data(sheet_obj, all_list):  # dosyadaki tüm verileri liste yapar
    for row in range(1,
                     sheet_obj.max_row + 1):  # row 1 den max satır sayısına kadar değer alıcak
        # Bu döngü tüm excel verilerini data_all boş listesine yazdırmak için
        temp = []  # Her bir satırı ayrı liste olarak all_list listesine
        # eklemek için geçici bir temp listesi oluşturduk
        # 1 eklememin sebebi satır ya da sütun seçerken son satır ya da son sütunu da seçebilmek
        for column in range(1, sheet_obj.max_column + 1):  # column 1 den max sütun sayısına kadar değer alıcak
            data_all = sheet_obj.cell(row=row,
                                      column=column).value  # exceldeki tüm değerler data_all variable olarak alınıyor
            temp.append(data_all)  # tüm değerleri ilk başta boş liste olarak tanımladığım temp e ekliyorum
        all_list.append(temp)  # Her bir satır liste olarak all_list listesine ekleniyor
    return all_list


def carrying_files(sheet_obj, ws2, wb2, filename):  #Bu fonksiyon excel dosyasındaki tüm verileri
    #boş excel dosyasına taşımak için
    for i in range(1, sheet_obj.max_row + 1):
        for j in range(1, sheet_obj.max_column + 1):
            # reading cell value from source excel file
            c = sheet_obj.cell(row=i, column=j)

            # writing the read value to destination excel file
            ws2.cell(row=i, column=j).value = c.value
        wb2.save(filename)
    return ws2, wb2


def copy_data(start_row, end_row, start_column,
              end_column, data_selected, filename, ws,
              wb):  #verileri istediğin yerden itibaren kopyalar, baştan itibaren yapıştırır
    range_row = (end_row - start_row)
    range_column = (end_column - start_column)
    for x in range(0, range_row):
        for y in range(0, range_column):
            ws.cell(row=(x + 1), column=(y + 1)).value = data_selected[x][y]
    wb.save(filename)
    return ws, wb


def carrying_spesific_dimension(start_row, end_row, start_column, end_column,
                               copy_start_row, copy_start_column,
                                wb2, ws2, filename, data_selected):
    #verileri istediğin yerden itibaren kopyalar, istediğin yerden itibaren yapıştırır

    temp_copy_start_column = copy_start_column  #iç içe for döngüsünde ikinci for döngüsünden sonra diğer satıra geçiyor.
    #başlangıç konumunu tempe atayarak her satırı aynı sütundan başlatabiliyoruz.

    for x in range(start_row, end_row):  # x kullanıcının seçtiği başlangıç ve bitiş satırları arasında
        # değer alıyor(Bitiş değeri dahil değil),
        for y in range(start_column, end_column):  # y kullanıcının seçtiği başlangıç ve
            # bitiş sütunları arasında değer alıyor(Bitiş değeri dahil değil )
            ws2.cell(row=copy_start_row, column=copy_start_column).value = data_selected[x - 1][y - 1]  # burada x-1 ve y-1 dememin sebebi
            # biz kullanıcıdan exceldeki satır ve sütunları istedik.
            # Listede index 0 dan başlıyor excelde 1 den.
            copy_start_column += 1  #hücredeki değerleri yazarken bir sonraki sütuna geçiyor.

        copy_start_column = temp_copy_start_column #yeni satıra geçtiğimizde yazmaya başlangıç sütunundan başlıyor.
        copy_start_row += 1  #yeni satıra geçiyor.
    wb2.save(filename)

def move_cell(ws,data_selected):  #sütunu en son sütunun bir sağına taşımayı sağlar
    max = ws.max_column
    for x in range(1, ws.max_row + 1):  # x kullanıcının seçtiği başlangıç ve bitiş satırları arasında
        for y in range(1, 2):  # y kullanıcının seçtiği başlangıç ve
            # bitiş sütunları arasında değer alıyor(Bitiş değeri dahil değil )
            ws.cell(row=x, column=max + 1).value = data_selected[x - 1][y - 1]

def delete_col(ws2, wb2, filename, col_start, col_start_num):  #istediğin sütundan itibaren istediğin miktarda sütun silmeni sağlar
    ws2.delete_cols(col_start, col_start_num)
    wb2.save(filename)
    return ws2, wb2

def remove(sheet, row):  #satır silmeyi sağlar
    for cell in row:
        if cell.value != None:
            return

    sheet.delete_rows(row[0].row, 1)


def filter(user_list, sheet_obj, ws2, data):  #yazdığımız adı tüm dosyada arar, yanındaki sütuna gönderdiği değeri bastırır
    all_datas = []
    all_data(sheet_obj, all_datas)  #sheetdeki tüm verileri liste olarak alır
    for x in range(0, sheet_obj.max_row):
        for y in range(0, sheet_obj.max_column):
            for i in range (0, len(user_list)):  #user_list e bütün bir str olarak ulaşır, her bir indexini ayrı ayrı almaz(bunun için user_list[i] yazmalıydık
                if str(user_list) in str(all_datas[x][y]):  #eğer user_list ve all_data listesindeki veriler aynıysa
                    ws2.cell(row=(x + 1), column=(y + 2)).value = data  #belirtilen celle mainden gönderdiğimiz data yı yazdır
                    #row un x+1 olmasının sebebi excelde 1 den başlaması listede 0 dan , column un y+2 olmasının sebebi bir sonraki satıra yazdırmak istememiz
    return ws2

def filter_and_delete(user_list,sheet_obj):  #gönderilen sözcüğü arar ve o sözcüğün olduğu satırı siler
    all_datas = []
    all_data(sheet_obj, all_datas)
    for x in range(0, sheet_obj.max_row):
        for y in range(0, sheet_obj.max_column):
            for i in range(0, len(user_list)):
                if str(user_list) in str(all_datas[x][y]):
                    sheet_obj.delete_rows(idx=(x+1), amount=1)  #buraya kadar filter fonksiyonunun aynısı
                    # burda x+1 dememizim sebebi listede 0 dan excelde 1 den başlıyor olması


    return sheet_obj


"""def filter_and_add(data_base, sheet_obj, user_list):

    data_selected = []
    real_max_column = 0
    print(real_max_column)
    data_select(sheet_obj, 2, sheet_obj.max_row + 1, 6,
                7, data_selected)  # En sağdaki sütunun tamamı
    print(data_selected)

    all_datas = []
    all_data(sheet_obj, all_datas)  # sheetdeki tüm verileri liste olarak alır

    total_list = []
    for i in range(0, len(data_base)):
        total_list1 = []
        for j in range(0, len(user_list)):
            data = 0
            total_list1.append(data)
        total_list.append(total_list1)

    print(total_list)
    total_list[0][2] = 5
    print(total_list)

    for x in range(1, len(data_selected)-2):  #son iki satırı almaması için, parantez işareti olduğu için str den floata çeviremiyor
        for y in range(0, len(data_base)):
            for k in range(0, len(user_list)):
                #print(user_list[k], all_datas[x][2])

                if str(data_base[y]) in str(data_selected[x]) and str(user_list[k]) in str(all_datas[x][2]):   # eğer data_base data_selected ın içindeyse
                    print("user_list", user_list[k], "all_datas", all_datas[x][2])
                    total_list[y][k] = str(float(total_list[y][k]) + float(sheet_obj.cell(row=(x+1), column=5).value))
                    print("y,k",y,k)
                    print(total_list)
                  
                    print("************")
    a = sheet_obj.max_row
    sheet_obj.cell(row=(a + 2), column=2).value = str(user_list[0])
    sheet_obj.cell(row=(a + 2), column=3).value = str(user_list[1]) + '/' + str(user_list[2])
   
    print(total_list)

    for i in range(0, len(data_base)-1):
        x = sheet_obj.max_row
        for j in range(0, len(user_list)-1):
            sheet_obj.cell(row=(x + 1), column=1).value = str(data_base[i])
            if str(user_list[j]) == 'stoklu':
                sheet_obj.cell(row=(x+ 1), column=3).value = str(float(total_list[i][j]) + float(total_list[i][j+1]))
            else:
                sheet_obj.cell(row=(x + 1), column=2).value = str(float(total_list[i][j]))

    return sheet_obj"""



def data_select(sheet_obj, start_row, end_row, start_column,
                end_column, data_selected):  # başlangıç ve bitiş satır ve sütununu seçerek onları liste yapar.
    for i in range(start_row, end_row):
        temp = []  # Her bir satırı ayrı liste olarak data_selected listesine eklemek için geçici bir temp listesi oluşturduk
        for j in range(start_column, end_column):
            data = sheet_obj.cell(row=i, column=j).value
            temp.append(data)  # döngünün sonunda geçici temp listesine tüm satır eklenecek
        data_selected.append(
            temp)  # Bu data_main seçtiğimiz aralığı göstermek için, excel'e yazan kısımda kullanıldığı takdirde out of range hatası alınıyor,
        # çünkü liste orijinal halinden daha kısa kalıyor
    return data_selected

def data_base(sheet_obj):  #seçtiğimiz listenin içindeki verileri tekrar etmeden yeni bir listeye koyar. her bir veriden yalnızca bir tane olur(unique)
    data_selected = []
    data_select(sheet_obj, 2, sheet_obj.max_row + 1, sheet_obj.max_column,
                sheet_obj.max_column + 1, data_selected)
    data_base_list = []
    for x in data_selected:
        if x not in data_base_list:  #seçilen data_selected daki veriler eğer data_base listesinde yoksa data_base e ekle
            data_base_list.append(x)
    return data_base_list


def calculate_amount(data_base, sheet_obj):  #en sağdaki sütundaki her bir indexi data_base fonksiyonuyla al
    data_selected = []
    data_select(sheet_obj, 2, sheet_obj.max_row + 1, sheet_obj.max_column,
                sheet_obj.max_column + 1, data_selected)   #En sağdaki sütunun tamamı
    total_list = []
    total = 0
    for i in range(0, len(data_base)):
        total_list.append('0')  #total_list deki her bir indexe 0 dedik çünkü soraki aşamada bunları kullanarak toplama işlemi yapacağız
    for x in range(0, len(data_selected)):  #en sağdaki sütunun uzunluğuna kadar
        for y in range(0, len(data_base)):  #data_base in uzunluğuna kadar
            if str(data_base[y]) in str(data_selected[x]):  #eğer data_base data_selected ın içindeyse
                if sheet_obj.cell(row=(x + 2), column=5).value != None:  #eğer 5. sütundaki celller boş değilse
                    total_list[y] = str(float(total_list[y]) + float(sheet_obj.cell(row=(x + 2), column=5).value))
                    #total_listin y inci elemanı + exceldeki 5. kolon (x+2). sütundaki sayıların toplamı yeni total_list[y] olsun

    a = sheet_obj.max_row
    for i in range(0, len(data_base)):
        sheet_obj.cell(row=(a+1), column=(i+1)).value = str(data_base[i])  #data_base i bitiş satırından sonraki satıra sırayla bastırır
        sheet_obj.cell(row=(a+2), column=(i+1)).value = total_list[i]   #data_base dekilerinin toplamını(total_list) sırayla bastırır

    return sheet_obj


def pandalar(df):

    print(df)
    df['Tutar'] = df['Tutar'].astype(str)
    df['Gy'] = df['Gy'].astype(str)

    df.sort_values(by=["Gy"], inplace=True, ascending=True)  # bütün dosyayı sıralıyor
    df_rownum = df.shape[0]  # 1 sütun sayısı, 0 satır sayısını gösteriyor
    df_colnum = df.shape[1]  # 1 sütun sayısı, 0 satır sayısını gösteriyor

    uniqueValues = df['Gy'].unique()  # ayrık değerleri veriyor, 84000 ve diğerleri

    yapım_list = ['dışarı yapılan işler', 'stoklu', 'stoksuz']


    df_colnum += 1  #max sütununun bir fazlasından başla
    #df_rownum += 1

    col_list = df.index.tolist()  # indexi listeye çevir
    df_rownum = 0  #satırları 0. indexten yapıştırmaya başla

    for i in range(0, len(uniqueValues)):
        kedi = df.loc[df['Gy'] == str(uniqueValues[i])]   #uniqueValues[i] her part numarası(sadece bir kere alınıyor)
        for j in range(0, len(yapım_list)):

            if j == 0:
                seperated = kedi.loc[kedi["Yapım"] == yapım_list[j]]  #
                seperated['Tutar'] = seperated['Tutar'].astype(np.float64)  #
                toplam = seperated["Tutar"].sum()
                df.loc[col_list[df_rownum], df_colnum] = toplam

                df.loc[col_list[df_rownum], df_colnum + 1] = str(uniqueValues[i])
                df.loc[col_list[df_rownum], df_colnum + 2] = yapım_list[j]
                df_rownum = df_rownum + 1

            elif j == 1:
                seperated = kedi.loc[kedi["Yapım"] == yapım_list[j]]
                seperated['Tutar'] = seperated['Tutar'].astype(np.float64)
                toplam = seperated["Tutar"].sum()

                seperated1 = kedi.loc[kedi["Yapım"] == yapım_list[j+1]]
                seperated1['Tutar'] = seperated1['Tutar'].astype(np.float64)
                toplam1 = seperated1["Tutar"].sum()


                df.loc[col_list[df_rownum], df_colnum] = toplam + toplam1  #DÜZENLENDİ !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

                df.loc[col_list[df_rownum], df_colnum + 1] = str(uniqueValues[i])
                df.loc[col_list[df_rownum], df_colnum + 2] = yapım_list[j] + '/' +  yapım_list[j+1]  #DÜZENLENDİ !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
                df_rownum = df_rownum + 1

    return df


if __name__ == '__main__':
    path = r"C:\Users\yaren\OneDrive\Masaüstü\mali_yet.xlsx"
    wb1 = openpyxl.load_workbook(path)
    ws1 = wb1.worksheets[0]
    ws2 = wb1.worksheets[1]

    filename = "büt_NEW.xlsx"  # bu ada sahip boş bir ecxel dosyası oluştur. pathi yazdırmadığımız için bu dosyanın olduğu klasöre açıyor
    wb2 = openpyxl.Workbook()
    ws3 = wb2.create_sheet('Sayfa1')  # sayfa1 adlı yeni sheet oluştur
    ws4 = wb2.create_sheet('Sayfa2')  # sayfa2 adlı yeni sheet oluştur
    wb2.save(filename)

    all_list1 = []
    all_data(ws1, all_list1)  # birinci sayfayı liste yapıyor

    all_list2 = []
    all_data(ws2, all_list2)  # ikinci sayfayı liste yapıyor

    copy_data(1, ws1.max_row + 1, 1,
              ws1.max_column, all_list1, filename, ws3, wb2)  # yeni açtığımız excel dosyasına 1. sayfayı kopyalıyor

    copy_data(1, ws2.max_row + 1, 1,
              ws2.max_column, all_list2, filename, ws4, wb2)  # yeni açtığımız excel dosyasına 2. sayfayı kopyalıyor

    delete_col(ws3, wb2, filename, 1, 8)  # 1. sayfadaki sütun silme işlemleri
    delete_col(ws3, wb2, filename, 2, 3)
    delete_col(ws3, wb2, filename, 5, 1)

    delete_col(ws4, wb2, filename, 1, 3)  # 2. sayfadaki sütun silem işlemleri
    delete_col(ws4, wb2, filename, 4, 1)
    delete_col(ws4, wb2, filename, 5, 1)
    delete_col(ws4, wb2, filename, 6, 1)

    ws3.insert_cols(3)  # 1. sayfadaki 3. sütunun soluna boş sütun açar, son halinde boş sütun 3. sütunda olur
    ws4.insert_cols(3)  # 2. sayfadaki 3. sütunun soluna boş sütun açar, son halinde boş sütun 3. sütunda olur

    data_selected2 = []
    data_select(ws3, 1, ws3.max_row + 1, 1,
                2, data_selected2)  # excelin 1. sayfasının ilk sütununu liste olarak alır

    move_cell(ws3, data_selected2)  # ilk sütunu en son sütunun bir sağına kopyalamayı sağlar
    delete_col(ws3, wb2, filename, 1, 1)  # ilk sütunun aynısı en son sütundada olur, ilk sütunu siliyoruz
    ws3.insert_cols(1)  # ilk sütunun boş olarak kalmasını istediğimiz için en sola boş sütun ekliyoruz

    data_selected3 = []
    data_select(ws3, 1, ws3.max_row + 1, 1,
                ws3.max_column + 1, data_selected3)  # excelin 2. sayfasının ilk sütununu liste olarak alır

    filter('DYI', ws3, ws3, 'dışarı yapılan işler')
    filter('STZ', ws3, ws3, 'stoksuz')



    all_list3 = []
    all_data(ws4,
             all_list3)  # ikinci sayfadaki gerçek max_row u bulmak için ikinci sayfadaki tüm verileri liste olarak aldık

    range_row = 0
    for x in range(0, ws4.max_row):  # 2. sayfada boş satırlar çok uzun, bunu bir şekilde önlemezsek
        # stoklu yazısı olması gereken max satır sayısının da altında yazmaya devam ediyor
        if all_list3[x][
            0] != None:  # eğer 1. sütundaki sayı none a eşit değilse rangi arttır. range bizim asıl max_row umuz
            range_row += 1
        else:
            break
    for x in range(0, range_row):
        ws4.cell(row=(x + 1), column=3).value = 'stoklu'

    ws4.delete_rows(idx=1, amount=1)  # 2. sayfadaki ilk satırı sil çünkü dosyayı taşırken bu satıra gerek yok

    all_list4 = []
    all_data(ws4,
             all_list4)  # ikinci sayfayı birinci sayfaya taşımak için ikinci sayfadaki tüm verileri liste olarak aldık
    carrying_spesific_dimension(1, range_row, 1, ws4.max_column + 1,
                                ws3.max_row + 1, 1, wb2, ws3, filename,
                                all_list4)  # bu fonksiyonu kullanarak ikinci sayfadaki verileri
    # birinci sayfadaki verilerin bitti yerden itibaren yapıştırdık



    sheet_remove1 = wb2[
        'Sheet']  # exceli kendimiz oluştururken ilk sayfayı otomatik atıyor, ben sonradan excel dosyasını kendi oluşturcak şekilde değiştirdiğim için
    # sayfa isimlerini değiştirmek istemedim onun yerine kendi oluşturduğu sayfayı silsin dedim
    sheet_remove2 = wb2[
        'Sayfa2']  # taşıma işlemini yaptıktan sonra bu sayfaya ihiyacımız kalmadı, tüm veriler Sayfa1 de
    wb2.remove(sheet_remove1)
    wb2.remove(sheet_remove2)



    """# filter_and_delete fonksiyonuyla gönderilen data taranıyor o datanın olduğu satır siliniyor
    filter_and_delete('SU', ws3)
    filter_and_delete('DOĞALGAZ', ws3)
    filter_and_delete('ELEKTRİK', ws3)  # excelde son k harfi farklı bir tipte olduğu için bu satırı silmiyor"""


    data_base_list = data_base(ws3)  # tekrar etmeyen şekilde tüm gy numaralarını aldı.
    # calculate_amount(data_base(ws3), ws3)  #dosyamızın son halini kullanarak belirli işlemleri yaptı

    ws3.cell(row=1, column=3).value = 'Yapım'  # yazmıyordu
    wb2.save(filename)



    data = pd.read_excel(filename)
    df = pd.DataFrame(data)

    liste1 = df["Yapım"].tolist()  #YENİ !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    liste2 = df["TF Açıklaması"].tolist()
    for i in range(0, len(liste1)):  #YENİ !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        if (liste1[i] != 'dışarı yapılan işler') and (liste1[i] != 'stoklu') and (liste1[i] != 'stoksuz')  :  #YENİ !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
            df['Yapım'] = df['Yapım'].replace([None], 'stoksuz')  #YENİ !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    for i in range(0,len(liste2)):
        if (liste2[i] =="ELEKTRİK") or (liste2[i] =="SU") or (liste2[i] =="DOĞALGAZ"):
            df.loc[i,"Yapım"] = "Enerji"
            #df["Yapım"]=df["Yapım"].replace(["stoksuz"],"ENERJİ")







    print(df.to_string())
    pandalar(df)  # pandalar fonksiyonunu çağırıyor

    df['Tutar'] = df['Tutar'].astype(np.float64)  #YENİ !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!  Virgül yapmak için
    """df_son = df.groupby(['Yapım'])['Tutar'].sum()
    print(df_son)
    df_son.plot.pie()
    plt.show() 


    df2 = df.groupby(['Yapım'])['Tutar'].sum().plot(
        kind='pie', y='Maliyet', autopct='%1.0f%%')
    df2.axis('equal')
    plt.rcParams['figure.dpi'] = 360
    plt.tight_layout()
    plt.show()"""


    df.to_excel('maliyet_yeni.xlsx', sheet_name='new_sheet_name')  # bu dosyaya kaydediyor




